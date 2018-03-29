import logging

import openpyxl
from ExcelRobot.reader import ExcelReader
from ExcelRobot.utils import (BoolFormat, DataType, DateFormat, NumberFormat,
                              copy_file, del_file, excel_name2coord,
                              get_file_path, is_file, random_temp_file)
from xlutils.copy import copy
from xlwt import Workbook, easyxf

LOGGER = logging.getLogger(__name__)


class XlsWriter:
    def __init__(self, workbook=None):
        self.wwb = copy(workbook) if workbook else Workbook(encoding='utf8')
        # Init sheet with new file
        if not workbook:
            self.create_sheet('Sheet')

    def create_sheet(self, sheet_name):
        self.wwb.add_sheet(sheet_name)

    def write_to_cell(self, sheet_name, column, row, value, data_format=None):
        sheet = self.wwb.get_sheet(sheet_name)
        sheet.write(row, column, label=value, style=easyxf('', num_format_str=data_format))

    def save(self, new_path):
        self.wwb.save(new_path)


class XlsxWriter:
    def __init__(self, file_path, is_new=False):
        LOGGER.debug('Opening writeable file at %s', file_path)
        self.wwb = openpyxl.Workbook() if is_new else openpyxl.load_workbook(file_path, data_only=True)

    def create_sheet(self, sheet_name):
        self.wwb.create_sheet(title=sheet_name)

    def write_to_cell(self, sheet_name, column, row, value, data_format=None):
        sheet = self.wwb[sheet_name]
        cell = sheet.cell(row + 1, column + 1, value)
        if data_format:
            cell.number_format = data_format

    def save(self, new_path):
        self.wwb.save(new_path)


class ExcelWriter(ExcelReader):

    def __init__(self, file_path, new_path=None, override=False,
                 date_format=DateFormat(), number_format=NumberFormat(), bool_format=BoolFormat()):
        self.is_update = new_path is None or new_path == file_path
        self.new_path = None if self.is_update else get_file_path(new_path)
        self.override = override
        if self.new_path and is_file(self.new_path):
            if self.override:
                del_file(self.new_path)
            else:
                raise FileExistsError('File ' + self.new_path + ' already existed. Use `override=True` to force override file')
        try:
            super().__init__(file_path, date_format, number_format, bool_format)
            self.writer = XlsWriter(self.workbook) if self.is_xls else XlsxWriter(self.file_path)
            self.is_new = False
        except FileNotFoundError as _:
            LOGGER.debug('Create new Excel file in %s', self.file_path)
            self.writer = XlsWriter() if self.is_xls else XlsxWriter(self.file_path, is_new=True)
            self.is_new = True

    def save_excel(self):
        """
        Saves the Excel file
        """
        save_path = self.file_path if self.is_new else random_temp_file(ext=self.extension) if self.is_update else self.new_path
        LOGGER.debug('Save Excel file to %s', save_path)
        self.writer.save(save_path)
        if not self.is_new and self.is_update:
            self.workbook.release_resources()
            copy_file(save_path, self.file_path, True)

    def create_sheet(self, sheet_name):
        """
        Creates and appends new Excel worksheet using the new sheet name to the current workbook.
        """
        self.writer.create_sheet(sheet_name)

    def write_to_cell_by_name(self, sheet_name, cell_name, value, data_type=None):
        col, row = excel_name2coord(cell_name)
        self.write_to_cell(sheet_name, col, row, value, data_type)

    def write_to_cell(self, sheet_name, column, row, value, data_type=None):
        """
        Using the sheet name the value of the indicated cell is set to be the value given in the parameter.
        """
        dtype = DataType.parse_type(data_type)
        LOGGER.debug('Write To Sheet: %s - Col: %s - Row: %s', sheet_name, column, row)
        LOGGER.debug('Data Type: %s', dtype)
        LOGGER.debug('Value Type: %s', type(value))
        if DataType.is_bool(dtype, value):
            self.writer.write_to_cell(sheet_name, column, row, self.bool_format.parse(value))
        elif DataType.is_date(dtype, value):
            raw_value = self.date_format.parse(dtype, value)
            dformat = self.date_format.get_excel_format(dtype, value)
            LOGGER.debug('Write Date Value:')
            LOGGER.debug('Raw Type: %s', type(raw_value))
            LOGGER.debug(raw_value)
            LOGGER.debug('Excel Format: %s', dformat)
            LOGGER.debug('-' * 10)
            self.writer.write_to_cell(sheet_name, column, row, raw_value, dformat)
        elif DataType.is_number(dtype, value):
            raw_value = self.number_format.parse(dtype, value)
            nformat = self.number_format.get_excel_format(dtype)
            LOGGER.debug('Raw Type: %s', type(raw_value))
            LOGGER.debug(raw_value)
            LOGGER.debug('Excel Format: %s', nformat)
            LOGGER.debug('-' * 10)
            self.writer.write_to_cell(sheet_name, column, row, raw_value, nformat)
        else:
            self.writer.write_to_cell(sheet_name, column, row, str(value))

    def modify_cell_with(self, sheet_name, column, row, op, val):
        """
        Using the sheet name a cell is modified with the given operation and value.
        """
        pass
        # my_sheet_index = self.workbook.sheet_names().index(sheet_name)
        # cell = self.workbook.get_sheet(my_sheet_index).cell(int(row), int(column))
        # curval = cell.value
        # if cell.ctype is XL_CELL_NUMBER:
        #     self.workbook.sheets()
        #     if not self.copied_workbook:
        #         self.copied_workbook = copy(self.workbook)
        #     plain = easyxf('')
        #     modexpr = str(curval) + op + val
        #     self.copied_workbook.get_sheet(my_sheet_index).write(int(row), int(column), eval(modexpr), plain)
